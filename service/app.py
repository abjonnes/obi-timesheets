from datetime import datetime, time, timedelta

from flask import Flask, jsonify, request, Response
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook


app = Flask(__name__)
TIMESHEET_TEMPLATE = 'template.xlsx'


@app.route("/service", methods=["POST"])
def post():
    form = dict(request.form)
    name = form.pop("name")[0]
    start_date = datetime.strptime(form.pop("startDate")[0], "%d%m%y")
    bts = {
        datetime.strptime(k, "%d%m%y"): v[0]
        for k, v in form.items()
    }
    timesheet = create_timesheet(name, start_date - timedelta(days=1), bts)
    filename = f"Timesheet-{name.replace(' ', '-')}-{(start_date + timedelta(days=5)).strftime('%Y-%m-%d')}.xlsx"
    return Response(timesheet, mimetype="application/vnd.ms-excel", headers={"Content-Disposition": f"attachment; filename={filename}"})


def create_timesheet(name, sunday_date, bts):
    wb = load_workbook(TIMESHEET_TEMPLATE)
    ws = wb['Sheet1']
    ws.cell(6, 2, value=name)
    ws.cell(10, 1, value=sunday_date)
    for d in range(5):
        weekday = sunday_date + timedelta(days=d+1)
        if bts.get(weekday) == "holiday":
            ws.cell(d + 11, 3, value="")
            ws.cell(d + 11, 4, value="")
        elif bts.get(weekday) == "full":
            ws.cell(d + 11, 3, value="")
            ws.cell(d + 11, 4, value="")
            ws.cell(d + 11, 7, value=8)
        elif bts.get(weekday) == "am":
            ws.cell(d + 11, 3, value=time(13, 0))
            ws.cell(d + 11, 7, value=4)
        elif bts.get(weekday) == "pm":
            ws.cell(d + 11, 4, value=time(13, 0))
            ws.cell(d + 11, 7, value=4)
    return save_virtual_workbook(wb)
